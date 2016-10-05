#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from .. import db
from flask_login import UserMixin
from flask import current_app


class Location(UserMixin, db.Model):

    __tablename__ = "locations"

    id = db.Column(db.Integer, primary_key=True)
    # Gemeinde
    plz = db.Column(db.Integer, index=True)
    locality = db.Column(db.String)
    bfs_nr = db.Column(db.Integer, index=True)
    district_nr = db.Column(db.Integer, index=True)
    district = db.Column(db.String)
    canton_nr = db.Column(db.Integer, index=True)
    canton = db.Column(db.String)
    canton_code = db.Column(db.String)

    ads = db.relationship('Ad')

    @staticmethod
    def insert():
        # Before we can insert locations we have to insert cantons and districts
        from openpyxl import load_workbook
        filename = 'data/Raumgliederungen_Schweiz_2015.xlsx'
        wb = load_workbook(filename=filename)
        sheet_ranges = wb['Gemeinde']
        current_app.logger.info("Inserting all localities (This may take a while)")
        for r in sheet_ranges.rows[1:]:
            l = db.session.query(Location).filter(Location.plz == r[0].value, Location.bfs_nr ==r[1].value).first()
            if l is None:
                d = db.session.query(District).filter(District.id == r[4].value).first()
                if d is None:
                    current_app.logger.error("Could not find district with id: {}".format(r[4].value))
                    continue
                c = db.session.query(Canton).filter(Canton.id == (d.id // 100)).first()
                locality = Location(plz=r[0].value, locality=r[2].value, bfs_nr=r[1].value,
                                    district_nr=r[4].value, district=d.name,
                                    canton_nr=c.id, canton=c.name,
                                    canton_code=c.code)
                db.session.add(locality)

        db.session.commit()

    def __repr__(self):
        return "<Location(id={}, plz={}, locality={}, bfs_nr={}, district_nr {} distirct={}, canton={})>".format(self.id,
                                                                                                                 self.plz,
                                                                                                                 self.locality,
                                                                                                                 self.bfs_nr,
                                                                                                                 self.district_nr,
                                                                                                                 self.district,
                                                                                                                 self.canton)


class Canton(UserMixin, db.Model):
    """
    Helper class to get the correct cantons
    """
    __tablename__ = "cantons"

    id = db.Column(db.Integer, primary_key=True, nullable=False)
    name = db.Column(db.String)
    code = db.Column(db.String, index=True)

    @staticmethod
    def insert():
        from openpyxl import load_workbook
        filename = 'data/Raumgliederungen_Schweiz_2015.xlsx'
        wb = load_workbook(filename=filename)
        sheet_ranges = wb['Kantone']
        current_app.logger.info("Inserting 26 cantons")
        for r in sheet_ranges.rows[1:]:
            canton = Canton(id=r[0].value, name=r[1].value, code=r[2].value)
            c = db.session.query(Canton).filter(Canton.name == canton.name).first()
            if c is None:
                db.session.add(canton)
        db.session.commit()


class District(UserMixin, db.Model):
    """
    Helper class to get the correct cantons
    """
    __tablename__ = "districts"

    id = db.Column(db.Integer, primary_key=True, nullable=False)
    name = db.Column(db.String)

    @staticmethod
    def insert():
        import warnings
        from openpyxl import load_workbook
        filename = 'data/Raumgliederungen_Schweiz_2015.xlsx'
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=filename)
        sheet_ranges = wb['Bezirke']
        current_app.logger.info("Inserting 148 Districts")
        for r in sheet_ranges.rows[1:]:
            d = db.session.query(District).filter(District.id==r[0].value).first()
            if d is None:
                district = District(id=r[0].value, name=r[1].value)
                db.session.add(district)
        db.session.commit()

